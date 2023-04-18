from openpyxl.styles.borders import Border, Side
import logging
import openpyxl
from datetime import datetime
from config import Config as cfg
import pandas

BEGIN_ROW_PATTERN = 10
BORDER = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

class BusinessExcel:
    # def write_task_log(self, data, workbook_path, text, period=''):
    #     logging.info('Run add_data_last_row')
    #     logging.info(f'Открываем файл: {workbook_path}')
    #     workbook = openpyxl.load_workbook(workbook_path)
    #     logging.debug('workbook.active')
    #     worksheet = workbook.active
    #     row = self.end_row(workbook, 1)
    #     logging.info(f'Строка для записи: {str(row)}')
    #
    #     if text == 'new_task':
    #         worksheet.cell(row=row, column=1).value = datetime.now()
    #         worksheet.cell(row=row, column=2).value = data['header']['requestID']
    #         worksheet.cell(row=row, column=3).value = str(data)
    #         worksheet.cell(row=row, column=4).value = data['header']['subject']
    #         worksheet.cell(row=row, column=5).value = period
    #     elif text == 'first_run':
    #         worksheet.cell(row=row, column=1).value = datetime.now()
    #         worksheet.cell(row=row, column=5).value = period
    #         worksheet.cell(row=row, column=6).value = str(data)
    #         worksheet.cell(row=row, column=7).value = 'Выполнено'
    #         worksheet.cell(row=row, column=8).value = datetime.now()
    #     else:
    #         worksheet.cell(row=row - 1, column=1).value = datetime.now()
    #         worksheet.cell(row=row - 1, column=6).value = str(data)
    #         worksheet.cell(row=row - 1, column=7).value = text
    #         worksheet.cell(row=row - 1, column=8).value = datetime.now()
    #
    #     logging.info(f'Save: {workbook_path}')
    #     workbook.save(workbook_path)
    #     logging.info('Запись в task_log добавлена')
    #     return

    # def end_row(self, workbook, column):
    #     """
    #     Записываем в лог файл информацию
    #     :param workbook: объект файла excel
    #     :param column: индекс столбца
    #     :return :row - индекс строки
    #     """
    #     sheet = workbook.active
    #     row = 2
    #     while True:
    #         if sheet.cell(row=row, column=column).value:
    #             row += 1
    #             continue
    #         return row

    def write_task_log(self, data, workbook_path):
        df = pandas.read_excel(workbook_path)
        df = df.append(data, ignore_index=True)
        df.to_excel(workbook_path, index=False)
        logging.info('Запись в task_log добавлена')


if __name__ == '__main__':
    data = {
        'requestID': '12',
        'type_index': 'PPI',
        'period': '2022-01',
        'body': {
            'd': 345,
            'g': 125
        }
    }
    e = BusinessExcel()
    workbook = openpyxl.load_workbook(cfg.task_log)
    e.write_task_log(data, cfg.task_log, 'new_task')
